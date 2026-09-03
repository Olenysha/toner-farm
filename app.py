"""
Тонер-фарм — учёт тонеров/картриджей для IT-отдела.
Локальный Flask-сервер для внутренней сети. Авторизация — доменная (AD/LDAP).

Модули:
- db.py              — пути, схема SQLite, миграции, сид, бэкапы, хелперы
- snmp_monitor.py    — фоновый SNMP-опрос принтеров (pysnmp 7, asyncio)
- auth.py            — доменная авторизация и сессии
- config/constants.py — константы (цвета-слоты, SNMP, карты)
- config/seed.json   — демо-данные принтеров и штрихкодов
"""
import json
import os
from datetime import datetime, date, timedelta
from io import BytesIO

import qrcode
from flask import (Flask, jsonify, redirect, render_template, request,
                   send_file, send_from_directory, session, url_for)
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageDraw

import auth
import request_logger
from config.constants import ALLOWED_PLAN_EXT, SLOT_COLUMN
from db import (BASE_DIR, DATA_DIR, STATIC_DIR, QR_DIR, FLOOR_PLAN_PATH,
                get_db, close_db, init_db, maybe_backup, row_to_dict,
                get_barcode, toner_with_info, serialize_printer, now_str)
from snmp_monitor import start_snmp_polling

app = Flask(__name__)
# True, когда сервер запущен с TLS-сертификатом (выставляется в __main__);
# по флагу base.html решает, нужен ли редирект http → https
app.config.setdefault('HTTPS_ENABLED', False)
app.secret_key = auth.load_or_create_secret(DATA_DIR)
app.permanent_session_lifetime = timedelta(days=auth.SESSION_DAYS)

# Логирование всех HTTP-запросов в logs/access.log + SSE на /logs
request_logger.init_request_logging(app)

app.teardown_appcontext(close_db)


@app.context_processor
def inject_auth():
    """Текущий пользователь/роль во все шаблоны (шапка, скрытие Админа)."""
    u = auth.current_user()
    return {'auth_user': u, 'auth_role': u['role'] if u else None}


@app.before_request
def before_request():
    # защита: без логина — только страница входа (у неё inline-стили)
    public = {'login', 'static', 'static_from_root'}
    if request.endpoint in public:
        return
    user = auth.current_user()
    if not user:
        if request.path.startswith('/api/'):
            return jsonify({'error': 'Требуется вход', 'login': '/login'}), 401
        return redirect('/login?next=' + request.path)
    # роли: view — только чтение; изменения (POST/PUT/DELETE) и админка — edit
    if user['role'] != 'edit':
        mutating = request.method not in ('GET', 'HEAD', 'OPTIONS')
        admin_page = request.endpoint in ('admin', 'qr_print')
        if mutating or admin_page:
            if request.path.startswith('/api/') or mutating:
                return jsonify({'error': f'Недостаточно прав: нужна группа {auth.EDIT_GROUP}'}), 403
            return redirect('/')
    maybe_backup()


# ------------------------------------------------- План этажа (заглушка PIL)

def _log_op(db, toner_id, printer_id, otype, old_toner_id, ts):
    """Запись операции с подписью текущего пользователя (из сессии)."""
    u = auth.current_user()
    db.execute(
        "INSERT INTO operations (toner_id, printer_id, type, old_toner_id, timestamp, user_name)"
        " VALUES (?,?,?,?,?,?)",
        (toner_id, printer_id, otype, old_toner_id, ts, u['name'] if u else None))

def ensure_floor_plan():
    """Рисуем простой план-заглушку, если файла нет."""
    if os.path.exists(FLOOR_PLAN_PATH):
        return
    w, h = 1200, 800
    img = Image.new('RGB', (w, h), '#f4f1ea')
    d = ImageDraw.Draw(img)
    # сетка
    for x in range(0, w, 50):
        d.line([(x, 0), (x, h)], fill='#e3ded2')
    for y in range(0, h, 50):
        d.line([(0, y), (w, y)], fill='#e3ded2')
    # комнаты
    rooms = [
        ((60, 60, 420, 380), 'Бухгалтерия'),
        ((60, 430, 420, 740), 'Коридор'),
        ((470, 60, 780, 380), 'HR'),
        ((470, 430, 780, 740), 'Кухня'),
        ((830, 60, 1140, 380), 'Отдел продаж'),
        ((830, 430, 1140, 740), 'Серверная'),
    ]
    for (x1, y1, x2, y2), label in rooms:
        d.rectangle([x1, y1, x2, y2], outline='#8a8578', width=4)
        d.text(((x1 + x2) // 2 - 40, (y1 + y2) // 2 - 8), label, fill='#6b675d')
    d.text((30, 10), 'Тонер-фарм — план этажа (заглушка, замените в админке)', fill='#4a463e')
    img.save(FLOOR_PLAN_PATH)


# ------------------------------------------------------------------ Страницы

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Вход по доменной учётке. Пока без принудительной защиты остальных роутов."""
    if auth.current_user():
        return redirect(request.args.get('next') or '/')
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')
        result = auth.try_login(username, password)
        if result:
            role, canonical = result
            auth.login_session(canonical, role)
            return redirect(request.form.get('next') or '/')
        error = (f'Неверный логин/пароль, нет членства в группах '
                 f'{auth.VIEW_GROUP}/{auth.EDIT_GROUP} или домен недоступен')
    return render_template('login.html', error=error,
                           groups_hint=f'{auth.VIEW_GROUP} / {auth.EDIT_GROUP}')


@app.route('/logout')
def logout():
    session.pop('user', None)
    return redirect('/login')


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/scan')
def scan():
    return render_template('scan.html')


@app.route('/item/<ean>')
def item(ean):
    return render_template('item.html', ean=ean)


@app.route('/admin')
def admin():
    return render_template('admin.html')


@app.route('/history')
def history():
    return render_template('history.html')


@app.route('/stats')
def stats():
    return render_template('stats.html')


@app.route('/availability')
def availability():
    return render_template('availability.html')


@app.route('/admin/qr_print')
def qr_print():
    db = get_db()
    rows = db.execute('SELECT ean_13, model_name, color FROM barcode_map ORDER BY model_name').fetchall()
    items = []
    for r in rows:
        png = os.path.join(QR_DIR, f'{r["ean_13"]}.png')
        if not os.path.exists(png):
            make_qr(r['ean_13'])
        items.append({'ean_13': r['ean_13'], 'model_name': r['model_name'],
                      'color': r['color'], 'qr_url': url_for('static', filename=f'qr_codes/{r["ean_13"]}.png')})
    return render_template('qr_print.html', items=items)


# ------------------------------------------------------------------ JSON API

@app.route('/api/lookup/<ean>')
def api_lookup(ean):
    db = get_db()
    bc = get_barcode(db, ean)
    if not bc:
        return jsonify({'found': False, 'ean_13': ean})
    # EAN общий для всех тонеров модели+цвета: считаем остаток на складе,
    # отдельно находим установленный экземпляр (если есть)
    stock_rows = db.execute(
        "SELECT id FROM toners WHERE ean_13 = ? AND status = 'stock' ORDER BY id",
        (ean,)).fetchall()
    installed = db.execute(
        "SELECT * FROM toners WHERE ean_13 = ? AND status = 'installed' ORDER BY id DESC LIMIT 1",
        (ean,)).fetchone()
    resp = {'found': True, 'barcode': bc, 'toner': None, 'printer': None,
            'stock_count': len(stock_rows),
            'stock_toner_id': stock_rows[-1]['id'] if stock_rows else None}
    if installed:
        resp['toner'] = toner_with_info(db, installed)
        if installed['current_printer_id']:
            p = db.execute('SELECT * FROM printers WHERE id = ?',
                           (installed['current_printer_id'],)).fetchone()
            resp['printer'] = row_to_dict(p)
    return jsonify(resp)


@app.route('/api/lookup_by_id/<int:tid>')
def api_lookup_by_id(tid):
    """Тонер по id + его запись из barcode_map (для режима выбора на карте)."""
    db = get_db()
    toner = db.execute('SELECT * FROM toners WHERE id = ?', (tid,)).fetchone()
    if not toner:
        return jsonify({'error': 'Тонер не найден'}), 404
    return jsonify({'toner': toner_with_info(db, toner), 'barcode': get_barcode(db, toner['ean_13'])})


@app.route('/api/barcode_map', methods=['POST'])
def api_barcode_map_create():
    """Обучение нового EAN: запись в barcode_map + тонеры на склад (quantity шт.)."""
    data = request.get_json(force=True)
    ean = (data.get('ean_13') or '').strip()
    if not ean:
        return jsonify({'error': 'ean_13 обязателен'}), 400
    qty = int(data.get('quantity') or 1)
    if qty < 1:
        return jsonify({'error': 'Количество должно быть ≥ 1'}), 400
    compat = data.get('compatible_printers') or []
    if isinstance(compat, str):
        compat = [compat]
    db = get_db()
    db.execute(
        'INSERT OR REPLACE INTO barcode_map (ean_13, model_name, color, compatible_printers, page_yield) VALUES (?,?,?,?,?)',
        (ean, data.get('model_name'), data.get('color'),
         json.dumps(compat, ensure_ascii=False), data.get('page_yield')))
    _add_stock(db, ean, qty, data.get('enterprise_id'))
    db.commit()
    return jsonify({'ok': True, 'barcode': get_barcode(db, ean)})


def _add_stock(db, ean, qty, enterprise_id=None):
    """Приход qty штук на склад предприятия + запись в operations.

    enterprise_id не передан/не найден → склад первого предприятия.
    """
    eid = None
    if enterprise_id:
        row = db.execute('SELECT id FROM enterprises WHERE id = ?',
                         (enterprise_id,)).fetchone()
        eid = row['id'] if row else None
    if not eid:
        eid = db.execute('SELECT MIN(id) FROM enterprises').fetchone()[0]
    for _ in range(qty):
        cur = db.execute(
            "INSERT INTO toners (ean_13, status, enterprise_id) VALUES (?, 'stock', ?)",
            (ean, eid))
        _log_op(db, cur.lastrowid, None, 'stock_add', None, now_str())


@app.route('/api/stock/add', methods=['POST'])
def api_stock_add():
    """Приход на склад: qty штук по известному EAN (код общий для модели+цвета)."""
    data = request.get_json(force=True)
    ean = (data.get('ean_13') or '').strip()
    qty = int(data.get('quantity') or 0)
    if qty < 1:
        return jsonify({'error': 'Количество должно быть ≥ 1'}), 400
    db = get_db()
    if not get_barcode(db, ean):
        return jsonify({'error': 'EAN не найден в справочнике'}), 404
    _add_stock(db, ean, qty, data.get('enterprise_id'))
    db.commit()
    return jsonify({'ok': True, 'added': qty})


@app.route('/api/stock/deplete', methods=['POST'])
def api_stock_deplete():
    """Ручное списание qty штук со склада по EAN (списываются самые старые).

    С enterprise_id — списание только со склада этого предприятия.
    """
    data = request.get_json(force=True)
    ean = (data.get('ean_13') or '').strip()
    qty = int(data.get('quantity') or 0)
    if qty < 1:
        return jsonify({'error': 'Количество должно быть ≥ 1'}), 400
    db = get_db()
    sql = "SELECT id FROM toners WHERE ean_13 = ? AND status = 'stock'"
    params = [ean]
    if data.get('enterprise_id'):
        sql += ' AND enterprise_id = ?'
        params.append(data['enterprise_id'])
    sql += ' ORDER BY id'
    rows = db.execute(sql, params).fetchall()
    if len(rows) < qty:
        return jsonify({'error': f'На складе только {len(rows)} шт.'}), 400
    for r in rows[:qty]:
        db.execute("UPDATE toners SET status='depleted' WHERE id=?", (r['id'],))
        _log_op(db, r['id'], None, 'depleted', None, now_str())
    db.commit()
    return jsonify({'ok': True, 'depleted': qty})


@app.route('/api/install', methods=['POST'])
def api_install():
    """Установка тонера в принтер; старый тонер слота автоматически списывается."""
    data = request.get_json(force=True)
    toner_id = data.get('toner_id')
    printer_id = data.get('printer_id')
    db = get_db()
    toner = db.execute('SELECT * FROM toners WHERE id = ?', (toner_id,)).fetchone()
    printer = db.execute('SELECT * FROM printers WHERE id = ?', (printer_id,)).fetchone()
    if not toner or not printer:
        return jsonify({'error': 'Тонер или принтер не найден'}), 404
    if toner['status'] == 'installed':
        return jsonify({'error': 'Тонер уже установлен'}), 400
    bc = get_barcode(db, toner['ean_13'])
    if not bc or bc['color'] not in SLOT_COLUMN:
        return jsonify({'error': 'Неизвестный цвет тонера'}), 400
    col = SLOT_COLUMN[bc['color']]  # цвет всегда из БД, вручную не выбирается
    if printer['type'] == 'mono' and col != 'toner_bk_id':
        return jsonify({'error': 'В моно-принтер можно ставить только чёрный тонер'}), 400

    old_toner_id = printer[col]
    now = now_str()
    if old_toner_id:
        # старый тонер из этого слота → списан
        db.execute("UPDATE toners SET status='depleted', current_printer_id=NULL WHERE id=?",
                   (old_toner_id,))
        _log_op(db, old_toner_id, printer_id, 'auto_depleted', None, now)
    db.execute(
        "UPDATE toners SET status='installed', current_printer_id=?, installed_at=? WHERE id=?",
        (printer_id, now, toner_id))
    db.execute(f'UPDATE printers SET {col}=? WHERE id=?', (toner_id, printer_id))
    _log_op(db, toner_id, printer_id, 'install', old_toner_id, now)
    db.commit()
    p = db.execute('SELECT * FROM printers WHERE id = ?', (printer_id,)).fetchone()
    return jsonify({'ok': True, 'auto_depleted_id': old_toner_id,
                    'printer': serialize_printer(db, p)})


@app.route('/api/return', methods=['POST'])
def api_return():
    """Возврат тонера на склад, слот принтера освобождается."""
    data = request.get_json(force=True)
    toner_id = data.get('toner_id')
    db = get_db()
    toner = db.execute('SELECT * FROM toners WHERE id = ?', (toner_id,)).fetchone()
    if not toner:
        return jsonify({'error': 'Тонер не найден'}), 404
    if toner['status'] != 'installed':
        return jsonify({'error': 'Тонер не установлен в принтер'}), 400
    printer_id = toner['current_printer_id']
    if printer_id:
        for col in SLOT_COLUMN.values():
            db.execute(f'UPDATE printers SET {col}=NULL WHERE id=? AND {col}=?',
                       (printer_id, toner_id))
    db.execute(
        "UPDATE toners SET status='stock', current_printer_id=NULL, installed_at=NULL WHERE id=?",
        (toner_id,))
    _log_op(db, toner_id, printer_id, 'return', None, now_str())
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/printers', methods=['GET'])
def api_printers_list():
    db = get_db()
    fid = request.args.get('floor_id', type=int)
    eid = request.args.get('enterprise_id', type=int)
    if fid:
        rows = db.execute('SELECT * FROM printers WHERE floor_id = ? ORDER BY id', (fid,)).fetchall()
    elif eid:
        rows = db.execute(
            'SELECT p.* FROM printers p '
            'JOIN floor_plans fp ON fp.id = p.floor_id '
            'WHERE fp.enterprise_id = ? ORDER BY p.id', (eid,)).fetchall()
    else:
        rows = db.execute('SELECT * FROM printers ORDER BY id').fetchall()
    return jsonify([serialize_printer(db, p) for p in rows])


@app.route('/api/printers', methods=['POST'])
def api_printers_create():
    data = request.get_json(force=True)
    ptype = data.get('type') or 'mono'
    db = get_db()
    floor_id = data.get('floor_id')
    if not floor_id:
        floor_id = db.execute('SELECT MIN(id) FROM floor_plans').fetchone()[0]
    cur = db.execute(
        'INSERT INTO printers (name, model, type, slots_count, x, y, ip_address, web_url, floor_id) VALUES (?,?,?,?,?,?,?,?,?)',
        (data.get('name'), data.get('model'), ptype,
         4 if ptype == 'color' else 1,
         float(data.get('x') or 50), float(data.get('y') or 50),
         data.get('ip_address'), data.get('web_url'), floor_id))
    db.commit()
    p = db.execute('SELECT * FROM printers WHERE id = ?', (cur.lastrowid,)).fetchone()
    return jsonify(serialize_printer(db, p)), 201


@app.route('/api/printers/<int:pid>', methods=['PUT'])
def api_printers_update(pid):
    data = request.get_json(force=True)
    db = get_db()
    p = db.execute('SELECT * FROM printers WHERE id = ?', (pid,)).fetchone()
    if not p:
        return jsonify({'error': 'Принтер не найден'}), 404
    fields = {}
    for key in ('name', 'model', 'type', 'ip_address', 'web_url'):
        if key in data:
            fields[key] = data[key]
    if 'floor_id' in data:
        fields['floor_id'] = int(data['floor_id'])
    for key in ('x', 'y'):
        if key in data:
            fields[key] = float(data[key])
    if 'type' in fields:
        fields['slots_count'] = 4 if fields['type'] == 'color' else 1
    if fields:
        sets = ', '.join(f'{k}=?' for k in fields)
        db.execute(f'UPDATE printers SET {sets} WHERE id=?', (*fields.values(), pid))
        db.commit()
    p = db.execute('SELECT * FROM printers WHERE id = ?', (pid,)).fetchone()
    return jsonify(serialize_printer(db, p))


@app.route('/api/printers/<int:pid>', methods=['DELETE'])
def api_printers_delete(pid):
    db = get_db()
    db.execute('DELETE FROM printers WHERE id = ?', (pid,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/printers/import_web_urls', methods=['POST'])
def api_import_web_urls():
    """Импорт web_url из Excel: колонки IP и web."""
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'error': 'Файл не выбран'}), 400
    try:
        wb = load_workbook(f)
        ws = wb.active
        headers = {}
        for cell in ws[1]:
            if cell.value:
                headers[str(cell.value).strip().lower()] = cell.column - 1
        if 'ip' not in headers or 'web' not in headers:
            return jsonify({'error': 'В файле должны быть колонки IP и web'}), 400
        db = get_db()
        updated = 0
        skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            ip = str(row[headers['ip']] or '').strip()
            url = str(row[headers['web']] or '').strip()
            if not ip or not url:
                skipped += 1
                continue
            cur = db.execute('UPDATE printers SET web_url=? WHERE ip_address=?', (url, ip))
            if cur.rowcount:
                updated += cur.rowcount
        db.commit()
        return jsonify({'ok': True, 'updated': updated, 'skipped': skipped})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/printers/<int:pid>/available_toners')
def api_printer_available_toners(pid):
    """Совместимые тонеры, которые сейчас есть на складе (для замены с карты).

    Для каждого подходящего EAN отдаём id складского экземпляра
    (для /api/install) и остаток в штуках.
    """
    db = get_db()
    p = db.execute('SELECT * FROM printers WHERE id = ?', (pid,)).fetchone()
    if not p:
        return jsonify({'error': 'Принтер не найден'}), 404
    # предприятие принтера (через этаж) — показываем только его склад
    prow = db.execute(
        'SELECT fp.enterprise_id AS eid FROM printers pr '
        'JOIN floor_plans fp ON fp.id = pr.floor_id WHERE pr.id = ?', (pid,)).fetchone()
    ent_id = prow['eid'] if prow else None
    out = []
    for r in db.execute('SELECT ean_13 FROM barcode_map'):
        bc = get_barcode(db, r['ean_13'])
        if not bc or p['model'] not in (bc['compatible_printers'] or []):
            continue
        if p['type'] == 'mono' and bc['color'] != 'Black':
            continue
        sql = "SELECT id FROM toners WHERE ean_13 = ? AND status = 'stock'"
        params = [bc['ean_13']]
        if ent_id:
            sql += ' AND enterprise_id = ?'
            params.append(ent_id)
        sql += ' ORDER BY id'
        stock = db.execute(sql, params).fetchall()
        if not stock:
            continue
        out.append({'toner_id': stock[-1]['id'], 'ean_13': bc['ean_13'],
                    'model_name': bc['model_name'], 'color': bc['color'],
                    'stock_count': len(stock)})
    return jsonify(out)


@app.route('/api/printers/<int:pid>/snmp')
def api_printer_snmp(pid):
    """Последние SNMP-данные конкретного принтера."""
    db = get_db()
    row = db.execute(
        'SELECT * FROM snmp_readings WHERE printer_id = ? ORDER BY timestamp DESC LIMIT 1',
        (pid,)).fetchone()
    if not row:
        return jsonify({'error': 'Нет данных'}), 404
    d = dict(row)
    d['alerts'] = json.loads(d['alerts'] or '[]')
    d['raw_data'] = json.loads(d['raw_data'] or '{}')
    return jsonify(d)


@app.route('/api/printers/snmp/all')
def api_all_printers_snmp():
    """Последние SNMP-данные всех принтеров (для карты/дашборда)."""
    db = get_db()
    rows = db.execute('''
        SELECT r.*, p.name AS printer_name, p.type
        FROM snmp_readings r
        JOIN (
            SELECT printer_id, MAX(timestamp) AS max_ts
            FROM snmp_readings GROUP BY printer_id
        ) latest ON r.printer_id = latest.printer_id AND r.timestamp = latest.max_ts
        JOIN printers p ON p.id = r.printer_id
    ''').fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d['alerts'] = json.loads(d['alerts'] or '[]')
        d['raw_data'] = json.loads(d['raw_data'] or '{}')
        result.append(d)
    return jsonify(result)


@app.route('/api/agents/snmp_report', methods=['POST'])
def api_agent_snmp_report():
    """Приём SNMP-данных от удалённого агента (Android/Termux, другая сеть).

    Агент опрашивает принтер локально и шлёт сюда JSON. Требуется AGENT_TOKEN.
    """
    from config.constants import AGENT_TOKEN
    if not AGENT_TOKEN:
        return jsonify({'error': 'AGENT_TOKEN не настроен на сервере'}), 403
    data = request.get_json(force=True)
    if data.get('token') != AGENT_TOKEN:
        return jsonify({'error': 'Неверный токен'}), 401

    printer_id = data.get('printer_id')
    ip_address = (data.get('ip_address') or '').strip()
    db = get_db()
    p = None
    if printer_id:
        p = db.execute('SELECT * FROM printers WHERE id = ?', (printer_id,)).fetchone()
    if not p and ip_address:
        p = db.execute('SELECT * FROM printers WHERE ip_address = ?', (ip_address,)).fetchone()
    if not p:
        return jsonify({'error': 'Принтер не найден'}), 404

    def _int(v):
        try:
            return int(v) if v is not None else None
        except (ValueError, TypeError):
            return None

    alerts = data.get('alerts') or []
    if isinstance(alerts, str):
        try:
            alerts = json.loads(alerts)
        except ValueError:
            alerts = []
    raw_data = data.get('raw_data') or {}
    if isinstance(raw_data, str):
        try:
            raw_data = json.loads(raw_data)
        except ValueError:
            raw_data = {}

    db.execute(
        '''INSERT INTO snmp_readings
           (printer_id, timestamp, black_level, cyan_level, magenta_level, yellow_level,
            page_counter, status_text, alerts, raw_data)
           VALUES (?,?,?,?,?,?,?,?,?,?)''',
        (p['id'], now_str(), _int(data.get('black_level')), _int(data.get('cyan_level')),
         _int(data.get('magenta_level')), _int(data.get('yellow_level')),
         _int(data.get('page_counter')), data.get('status_text') or 'Unknown',
         json.dumps(alerts, ensure_ascii=False),
         json.dumps(raw_data, ensure_ascii=False)))
    db.commit()
    return jsonify({'ok': True, 'printer_id': p['id']})


@app.route('/api/barcode_map/<ean>', methods=['PUT'])
def api_barcode_map_update(ean):
    data = request.get_json(force=True)
    db = get_db()
    if not db.execute('SELECT 1 FROM barcode_map WHERE ean_13=?', (ean,)).fetchone():
        return jsonify({'error': 'EAN не найден'}), 404
    compat = data.get('compatible_printers') or []
    if isinstance(compat, str):
        compat = [compat]
    db.execute(
        'UPDATE barcode_map SET model_name=?, color=?, compatible_printers=?, page_yield=? WHERE ean_13=?',
        (data.get('model_name'), data.get('color'),
         json.dumps(compat, ensure_ascii=False), data.get('page_yield'), ean))
    db.commit()
    return jsonify({'ok': True, 'barcode': get_barcode(db, ean)})


@app.route('/api/barcode_map/<ean>', methods=['DELETE'])
def api_barcode_map_delete(ean):
    db = get_db()
    db.execute('DELETE FROM barcode_map WHERE ean_13 = ?', (ean,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/barcode_map', methods=['GET'])
def api_barcode_map_list():
    db = get_db()
    rows = db.execute('SELECT * FROM barcode_map ORDER BY model_name').fetchall()
    return jsonify([get_barcode(db, r['ean_13']) for r in rows])


# ALLOWED_PLAN_EXT импортирован из config.constants


def plan_json(row):
    """Этаж → JSON с URL картинки."""
    rel = row['image_path']
    if rel.startswith('static/'):
        rel = rel[len('static/'):]
    return {'id': row['id'], 'name': row['name'],
            'enterprise_id': row['enterprise_id'],
            'image_url': url_for('static_from_root', path=rel)}


# ---------------------------------------------------------------- Предприятия

@app.route('/api/enterprises', methods=['GET'])
def api_enterprises_list():
    db = get_db()
    rows = db.execute('SELECT * FROM enterprises ORDER BY id').fetchall()
    return jsonify([{'id': r['id'], 'name': r['name']} for r in rows])


@app.route('/api/enterprises', methods=['POST'])
def api_enterprises_create():
    data = request.get_json(force=True)
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Название обязательно'}), 400
    db = get_db()
    cur = db.execute('INSERT INTO enterprises (name) VALUES (?)', (name,))
    db.commit()
    return jsonify({'id': cur.lastrowid, 'name': name}), 201


@app.route('/api/enterprises/<int:eid>', methods=['PUT'])
def api_enterprises_update(eid):
    data = request.get_json(force=True)
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Название обязательно'}), 400
    db = get_db()
    if not db.execute('SELECT 1 FROM enterprises WHERE id=?', (eid,)).fetchone():
        return jsonify({'error': 'Предприятие не найдено'}), 404
    db.execute('UPDATE enterprises SET name=? WHERE id=?', (name, eid))
    db.commit()
    return jsonify({'id': eid, 'name': name})


@app.route('/api/enterprises/<int:eid>', methods=['DELETE'])
def api_enterprises_delete(eid):
    db = get_db()
    if not db.execute('SELECT 1 FROM enterprises WHERE id=?', (eid,)).fetchone():
        return jsonify({'error': 'Предприятие не найдено'}), 404
    if db.execute('SELECT 1 FROM floor_plans WHERE enterprise_id=?', (eid,)).fetchone():
        return jsonify({'error': 'У предприятия есть этажи — сначала удалите их'}), 400
    db.execute('DELETE FROM enterprises WHERE id=?', (eid,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/floor_plans', methods=['GET'])
def api_floor_plans_list():
    """Список этажей; ?enterprise_id= фильтрует по предприятию."""
    db = get_db()
    eid = request.args.get('enterprise_id', type=int)
    if eid:
        rows = db.execute(
            'SELECT * FROM floor_plans WHERE enterprise_id=? ORDER BY id', (eid,)).fetchall()
    else:
        rows = db.execute('SELECT * FROM floor_plans ORDER BY id').fetchall()
    return jsonify([plan_json(r) for r in rows])


@app.route('/api/floor_plans', methods=['POST'])
def api_floor_plans_create():
    """Новый этаж у предприятия (план-заглушка, заменяется загрузкой файла)."""
    data = request.get_json(force=True)
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Название этажа обязательно'}), 400
    db = get_db()
    eid = data.get('enterprise_id')
    if eid:
        if not db.execute('SELECT 1 FROM enterprises WHERE id=?', (eid,)).fetchone():
            return jsonify({'error': 'Предприятие не найдено'}), 404
    else:
        eid = db.execute('SELECT MIN(id) FROM enterprises').fetchone()[0]
    cur = db.execute(
        'INSERT INTO floor_plans (name, image_path, enterprise_id) VALUES (?,?,?)',
        (name, 'static/floor_plan.png', eid))
    db.commit()
    row = db.execute('SELECT * FROM floor_plans WHERE id=?', (cur.lastrowid,)).fetchone()
    return jsonify(plan_json(row)), 201


@app.route('/api/floor_plans/<int:fid>', methods=['PUT'])
def api_floor_plans_update(fid):
    """Переименование этажа."""
    data = request.get_json(force=True)
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Название этажа обязательно'}), 400
    db = get_db()
    if not db.execute('SELECT 1 FROM floor_plans WHERE id=?', (fid,)).fetchone():
        return jsonify({'error': 'Этаж не найден'}), 404
    db.execute('UPDATE floor_plans SET name=? WHERE id=?', (name, fid))
    db.commit()
    row = db.execute('SELECT * FROM floor_plans WHERE id=?', (fid,)).fetchone()
    return jsonify(plan_json(row))


@app.route('/api/floor_plans/<int:fid>', methods=['DELETE'])
def api_floor_plans_delete(fid):
    db = get_db()
    if not db.execute('SELECT 1 FROM floor_plans WHERE id=?', (fid,)).fetchone():
        return jsonify({'error': 'Этаж не найден'}), 404
    if db.execute('SELECT 1 FROM printers WHERE floor_id=?', (fid,)).fetchone():
        return jsonify({'error': 'На этаже есть принтеры — сначала удалите или перенесите их'}), 400
    db.execute('DELETE FROM floor_plans WHERE id=?', (fid,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/floor_plan', methods=['POST'])
def api_floor_plan_upload():
    """Загрузка плана (JPEG/PNG) на конкретный этаж (поле floor_id)."""
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'error': 'Файл не выбран'}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ALLOWED_PLAN_EXT:
        return jsonify({'error': 'Нужен JPEG или PNG файл'}), 400
    db = get_db()
    fid = request.form.get('floor_id', type=int)
    if fid:
        plan = db.execute('SELECT * FROM floor_plans WHERE id = ?', (fid,)).fetchone()
    else:
        plan = db.execute('SELECT * FROM floor_plans ORDER BY id LIMIT 1').fetchone()
    if not plan:
        return jsonify({'error': 'Этаж не найден'}), 404
    rel = f'static/floor_plan_{plan["id"]}_{datetime.now().strftime("%Y%m%d%H%M%S")}{ext}'
    f.save(os.path.join(BASE_DIR, rel))
    db.execute('UPDATE floor_plans SET image_path = ? WHERE id = ?', (rel, plan['id']))
    db.commit()
    row = db.execute('SELECT * FROM floor_plans WHERE id = ?', (plan['id'],)).fetchone()
    resp = plan_json(row)
    resp['ok'] = True
    resp['image_path'] = resp['image_url']  # совместимость со старым фронтом
    return jsonify(resp)


# отдача файлов из static/ (в т.ч. загруженных планов)
@app.route('/static/<path:path>')
def static_from_root(path):
    return send_from_directory(STATIC_DIR, path)


@app.route('/api/floor_plan', methods=['GET'])
def api_floor_plan_active():
    """План конкретного этажа (?floor_id=), по умолчанию — первый."""
    db = get_db()
    fid = request.args.get('floor_id', type=int)
    if fid:
        row = db.execute('SELECT * FROM floor_plans WHERE id = ?', (fid,)).fetchone()
    else:
        row = db.execute('SELECT * FROM floor_plans ORDER BY id LIMIT 1').fetchone()
    if not row:
        return jsonify({'image_url': url_for('static', filename='floor_plan.png')})
    return jsonify(plan_json(row))


def make_qr(ean):
    """Генерация PNG QR-кода, кодирующего EAN."""
    os.makedirs(QR_DIR, exist_ok=True)
    path = os.path.join(QR_DIR, f'{ean}.png')
    img = qrcode.make(ean)
    img.save(path)
    return path


@app.route('/api/qr/<ean>')
def api_qr(ean):
    make_qr(ean)
    return jsonify({'ok': True, 'url': url_for('static_from_root', path=f'qr_codes/{ean}.png')})


@app.route('/api/stats')
def api_stats():
    db = get_db()
    eid = request.args.get('enterprise_id', type=int)
    # при фильтре по предприятию учитываем только операции его принтеров
    ent_join = ''
    ent_where = ''
    params = []
    if eid:
        ent_join = (' JOIN printers pe ON pe.id = o.printer_id'
                    ' JOIN floor_plans fe ON fe.id = pe.floor_id')
        ent_where = ' AND fe.enterprise_id = ?'
        params.append(eid)
    # замены по месяцам
    per_month = db.execute(
        "SELECT strftime('%Y-%m', o.timestamp) AS m, COUNT(*) AS c FROM operations o"
        + ent_join + " WHERE o.type='install'" + ent_where +
        " GROUP BY m ORDER BY m", params).fetchall()
    # средний срок жизни тонера по моделям (install -> auto_depleted)
    # v2: предиктивная аналитика по page_yield (остаток ресурса, прогноз даты замены)
    lifespan_join = ''
    lifespan_where = ''
    lifespan_params = []
    if eid:
        lifespan_join = (' JOIN printers pe ON pe.id = i.printer_id'
                         ' JOIN floor_plans fe ON fe.id = pe.floor_id')
        lifespan_where = ' AND fe.enterprise_id = ?'
        lifespan_params.append(eid)
    lifespan = db.execute(
        """
        SELECT bm.model_name, AVG(julianday(d.timestamp) - julianday(i.timestamp)) AS avg_days, COUNT(*) AS n
        FROM operations d
        JOIN operations i ON i.toner_id = d.toner_id AND i.type = 'install'
        JOIN toners t ON t.id = d.toner_id
        JOIN barcode_map bm ON bm.ean_13 = t.ean_13
        """ + lifespan_join + "\nWHERE d.type = 'auto_depleted'" + lifespan_where +
        "\nGROUP BY bm.model_name", lifespan_params).fetchall()
    # топ принтеров по числу замен
    top_printers = db.execute(
        "SELECT p.name, COUNT(*) AS c FROM operations o JOIN printers p ON p.id = o.printer_id "
        + ('JOIN floor_plans fp ON fp.id = p.floor_id ' if eid else '')
        + "WHERE o.type='install'" + (' AND fp.enterprise_id = ?' if eid else '') +
        " GROUP BY p.name ORDER BY c DESC LIMIT 10", params if eid else []).fetchall()
    # склад по моделям — общий, от предприятия не зависит
    stock = db.execute(
        "SELECT bm.model_name, bm.color, COUNT(*) AS c FROM toners t "
        "JOIN barcode_map bm ON bm.ean_13 = t.ean_13 "
        "WHERE t.status='stock' GROUP BY bm.model_name, bm.color ORDER BY bm.model_name").fetchall()
    return jsonify({
        'replacements_per_month': [{'month': r['m'], 'count': r['c']} for r in per_month],
        'lifespan_by_model': [{'model': r['model_name'], 'avg_days': round(r['avg_days'], 1) if r['avg_days'] else 0,
                               'samples': r['n']} for r in lifespan],
        'top_printers': [{'name': r['name'], 'count': r['c']} for r in top_printers],
        'stock_by_model': [{'model': r['model_name'], 'color': r['color'], 'count': r['c']} for r in stock],
    })


def query_history(args):
    db = get_db()
    sql = (
        "SELECT o.*, t.ean_13, bm.model_name, bm.color, p.name AS printer_name "
        "FROM operations o "
        "LEFT JOIN toners t ON t.id = o.toner_id "
        "LEFT JOIN barcode_map bm ON bm.ean_13 = t.ean_13 "
        "LEFT JOIN printers p ON p.id = o.printer_id WHERE 1=1")
    params = []
    if args.get('enterprise_id', type=int):
        # операции только принтеров выбранного предприятия (по этажам);
        # складские операции без принтера при фильтре скрываются — склад общий
        sql += (' AND o.printer_id IN (SELECT p2.id FROM printers p2 '
                'JOIN floor_plans fp2 ON fp2.id = p2.floor_id '
                'WHERE fp2.enterprise_id = ?)')
        params.append(args.get('enterprise_id', type=int))
    if args.get('date_from'):
        sql += ' AND o.timestamp >= ?'
        params.append(args['date_from'] + ' 00:00:00')
    if args.get('date_to'):
        sql += ' AND o.timestamp <= ?'
        params.append(args['date_to'] + ' 23:59:59')
    if args.get('printer_id'):
        sql += ' AND o.printer_id = ?'
        params.append(args['printer_id'])
    if args.get('type'):
        sql += ' AND o.type = ?'
        params.append(args['type'])
    sql += ' ORDER BY o.timestamp DESC, o.id DESC'
    return db.execute(sql, params).fetchall()


@app.route('/api/history')
def api_history():
    rows = query_history(request.args)
    return jsonify([dict(r) for r in rows])


@app.route('/history/export')
def history_export():
    rows = query_history(request.args)
    wb = Workbook()
    ws = wb.active
    ws.title = 'История операций'
    ws.append(['Дата/время', 'Тип', 'Модель тонера', 'EAN-13', 'Цвет',
               'Принтер', 'ID старого тонера', 'Пользователь'])
    type_ru = {'install': 'Установка', 'return': 'Возврат на склад',
               'auto_depleted': 'Авто-списание', 'stock_add': 'Приход на склад',
               'depleted': 'Списание'}
    for r in rows:
        ws.append([r['timestamp'], type_ru.get(r['type'], r['type']),
                   r['model_name'], r['ean_13'], r['color'],
                   r['printer_name'], r['old_toner_id'], r['user_name']])
    for col_cells in ws.columns:
        width = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells) + 2
        ws.column_dimensions[col_cells[0].column_letter].width = min(width, 40)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'toner_history_{date.today().isoformat()}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/availability')
def api_availability():
    """Наличие тонеров на складе сгруппировано по моделям принтеров.

    Принтеры с разными именами, но одинаковой моделью объединяются в одну
    группу. Для каждого цвета суммируются остатки всех EAN (в т.ч. аналогов
    других производителей), у которых модель принтера есть в compatible_printers.
    """
    db = get_db()
    eid = request.args.get('enterprise_id', type=int)
    stock_sql = "SELECT ean_13, COUNT(*) AS c FROM toners WHERE status='stock'"
    stock_params = []
    if eid:
        stock_sql += ' AND enterprise_id = ?'
        stock_params.append(eid)
    stock_sql += ' GROUP BY ean_13'
    stock_by_ean = {r['ean_13']: r['c'] for r in db.execute(stock_sql, stock_params)}
    bcs = [get_barcode(db, r['ean_13'])
           for r in db.execute('SELECT ean_13 FROM barcode_map')]
    if eid:
        printers = db.execute(
            'SELECT p.* FROM printers p JOIN floor_plans fp ON fp.id = p.floor_id '
            'WHERE fp.enterprise_id = ? ORDER BY p.name', (eid,)).fetchall()
    else:
        printers = db.execute('SELECT * FROM printers ORDER BY name').fetchall()
    groups = {}
    for p in printers:
        g = groups.setdefault(p['model'], {
            'model': p['model'], 'type': p['type'], 'printers': [], 'colors': {}})
        g['printers'].append(p['name'])
        if p['type'] == 'color':
            g['type'] = 'color'
    for g in groups.values():
        needed = ['Black'] if g['type'] == 'mono' else list(SLOT_COLUMN)
        colors = {c: 0 for c in needed}
        for bc in bcs:
            if not bc or bc['color'] not in colors:
                continue
            if g['model'] in (bc['compatible_printers'] or []):
                colors[bc['color']] += stock_by_ean.get(bc['ean_13'], 0)
        g['colors'] = colors
    return jsonify(sorted(groups.values(), key=lambda g: g['model'] or ''))


def _stock_rows(db, enterprise_id=None):
    """Остатки по моделям/цветам; с enterprise_id — только склад этого предприятия."""
    sql = ("SELECT bm.model_name, bm.color, COUNT(*) AS c FROM toners t "
           "JOIN barcode_map bm ON bm.ean_13 = t.ean_13 "
           "WHERE t.status='stock'")
    params = []
    if enterprise_id:
        sql += ' AND t.enterprise_id = ?'
        params.append(enterprise_id)
    sql += ' GROUP BY bm.model_name, bm.color ORDER BY bm.model_name'
    return db.execute(sql, params).fetchall()


@app.route('/api/stock')
def api_stock():
    db = get_db()
    eid = request.args.get('enterprise_id', type=int)
    rows = _stock_rows(db, eid)
    return jsonify([{'model': r['model_name'], 'color': r['color'], 'count': r['c']} for r in rows])


@app.route('/stock/export')
def stock_export():
    db = get_db()
    eid = request.args.get('enterprise_id', type=int)
    rows = _stock_rows(db, eid)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Остатки на складе'
    ws.append(['Модель тонера', 'Цвет', 'Количество, шт.'])
    for r in rows:
        ws.append([r['model_name'], r['color'], r['c']])
    for col_cells in ws.columns:
        width = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells) + 2
        ws.column_dimensions[col_cells[0].column_letter].width = min(width, 40)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'stock_{date.today().isoformat()}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ------------------------------------------------------------------ Запуск

init_db()
ensure_floor_plan()

if __name__ == '__main__':
    # Сертификат: явные пути из env (Docker монтирует ./certs) или ./certs/*.pem
    cert = os.environ.get('CERT_FILE') or os.path.join(BASE_DIR, 'certs', 'cert.pem')
    key = os.environ.get('KEY_FILE') or os.path.join(BASE_DIR, 'certs', 'key.pem')
    if os.path.exists(cert) and os.path.exists(key):
        ssl_ctx = (cert, key)
        app.config['HTTPS_ENABLED'] = True
        print('[HTTPS] https://<host>:5000')
    else:
        ssl_ctx = None
        app.config['HTTPS_ENABLED'] = False
        print('[HTTP]  http://<host>:5000 (сканер не заработает без HTTPS — используй reverse proxy или certs/)')

    # Фоновый SNMP-опрос принтеров (в Docker — отдельный процесс snmp_poller.py)
    start_snmp_polling()

    app.run(host='0.0.0.0', port=5000, ssl_context=ssl_ctx, threaded=True)